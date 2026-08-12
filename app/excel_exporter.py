from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import TextUnit, ValidatedSuggestion

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
STATIC_FONT = Font(color="666666")
IMPORTED_FONT = Font(color="008000")
REVIEW_FILL = PatternFill("solid", fgColor="FCE4D6")

SUGGESTION_HEADERS = [
    "Förslags-ID", "Dokument", "Kapitel", "Vers", "Hänvisning", "PDF-sida",
    "Ursprunglig text", "Föreslagen text", "Feltyp", "Motivering", "Kontext",
    "Säkerhet", "Status",
]
SUGGESTION_WIDTHS = {
    1: 12, 2: 24, 3: 9, 4: 8, 5: 28, 6: 10, 7: 35, 8: 35,
    9: 28, 10: 42, 11: 70, 12: 11, 13: 18,
}


def _style_sheet(ws, widths: dict[int, int]) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32
    for index, width in widths.items():
        ws.column_dimensions[get_column_letter(index)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _append_suggestion_sheet(ws, suggestions: list[ValidatedSuggestion]) -> None:
    ws.append(SUGGESTION_HEADERS)
    for item in suggestions:
        ws.append([
            item.suggestion_id, item.document, item.chapter, item.verse, item.reference, item.page,
            item.old, item.new, item.error_type, item.motivation, item.original_context,
            item.confidence, item.status,
        ])
    _style_sheet(ws, SUGGESTION_WIDTHS)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = IMPORTED_FONT
        row[12].fill = REVIEW_FILL


def _safe_sheet_name(name: str, used_names: set[str]) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", "-", name).strip().strip("'")
    cleaned = re.sub(r"\s+", " ", cleaned) or "Dokument"
    base = cleaned[:31]
    candidate = base
    counter = 2
    lowered = {item.lower() for item in used_names}
    while candidate.lower() in lowered:
        suffix = f" ({counter})"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        counter += 1
    used_names.add(candidate)
    return candidate


def export_workbook(
    output_path: Path,
    suggestions: list[ValidatedSuggestion],
    units: list[TextUnit],
    rejected: list[dict],
    diagnostics: dict,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ändringsförslag"
    _append_suggestion_sheet(ws, suggestions)

    units_ws = wb.create_sheet("Extraherad text")
    units_ws.append([
        "Unit-ID", "Dokument", "Kapitel", "Vers", "Hänvisning", "Vers infererad", "PDF-sida start",
        "PDF-sida slut", "Tryckt sida", "Spalt", "Rad start", "Rad slut", "Text", "Källrader",
    ])
    for unit in units:
        units_ws.append([
            unit.unit_id, unit.document, unit.chapter, unit.verse, unit.reference,
            "ja" if unit.verse_inferred else "nej", unit.page_start, unit.page_end,
            unit.printed_page_start, unit.column_start, unit.line_start, unit.line_end,
            unit.text, "\n".join(unit.source_lines),
        ])
    _style_sheet(units_ws, {1: 14, 2: 18, 3: 9, 4: 8, 5: 22, 6: 14, 7: 13, 8: 13,
                            9: 12, 10: 11, 11: 10, 12: 10, 13: 90, 14: 70})
    for row in units_ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = IMPORTED_FONT

    diag_ws = wb.create_sheet("Körinformation")
    diag_ws.append(["Fält", "Värde"])
    for key, value in diagnostics.items():
        diag_ws.append([key, json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value])
    diag_ws.append(["godkända_förslag", len(suggestions)])
    diag_ws.append(["avvisade_förslag", len(rejected)])
    _style_sheet(diag_ws, {1: 30, 2: 110})
    for row in diag_ws.iter_rows(min_row=2):
        row[0].font = STATIC_FONT
        row[1].font = IMPORTED_FONT

    rejected_ws = wb.create_sheet("Avvisade förslag")
    rejected_ws.append(["Unit-ID", "Original", "Förslag", "Orsaker", "Rådata"])
    for item in rejected:
        suggestion = item.get("suggestion", {})
        rejected_ws.append([
            suggestion.get("unit_id"), suggestion.get("old"), suggestion.get("new"),
            ", ".join(item.get("reasons", [])), json.dumps(item, ensure_ascii=False),
        ])
    _style_sheet(rejected_ws, {1: 14, 2: 35, 3: 35, 4: 36, 5: 90})

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def export_batch_workbook(output_path: Path, results: list[dict], batch_diagnostics: dict) -> None:
    """Create one workbook with a summary and one suggestion sheet per successful document."""
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Sammanställning"
    summary_ws.append([
        "Dokument", "Filnamn", "Status", "Ändringsförslag", "Textenheter",
        "GPT-anrop", "Tid (s)", "Felmeddelande",
    ])

    for result in results:
        diagnostics = result.get("diagnostics") or {}
        summary_ws.append([
            result.get("document") or Path(result["filename"]).stem,
            result["filename"],
            "Klar" if result.get("success") else "Misslyckades",
            diagnostics.get("accepted_suggestion_count", 0),
            diagnostics.get("unit_count", 0),
            diagnostics.get("estimated_gpt_calls", 0),
            diagnostics.get("total_seconds", ""),
            result.get("error", ""),
        ])

    _style_sheet(summary_ws, {1: 36, 2: 36, 3: 16, 4: 18, 5: 14, 6: 12, 7: 12, 8: 80})
    for row in summary_ws.iter_rows(min_row=2):
        row[0].font = IMPORTED_FONT
        row[1].font = IMPORTED_FONT
        row[2].font = IMPORTED_FONT

    used_names = {"Sammanställning"}
    for result in results:
        if not result.get("success"):
            continue
        sheet_name = _safe_sheet_name(result.get("document") or Path(result["filename"]).stem, used_names)
        ws = wb.create_sheet(sheet_name)
        _append_suggestion_sheet(ws, result["suggestions"])

    info_ws = wb.create_sheet("Körinformation")
    info_ws.append(["Fält", "Värde"])
    for key, value in batch_diagnostics.items():
        info_ws.append([key, json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value])
    _style_sheet(info_ws, {1: 30, 2: 110})
    for row in info_ws.iter_rows(min_row=2):
        row[0].font = STATIC_FONT
        row[1].font = IMPORTED_FONT

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)